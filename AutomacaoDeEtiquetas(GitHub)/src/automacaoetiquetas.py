import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# ==========================================================
# CONFIGURAÇÕES
# ==========================================================

BASE_DIR = os.getenv("AUTOMACAO_ETIQUETAS_BASE_DIR", os.path.dirname(os.path.abspath(__file__)))
PASTA_ENTRADA = os.path.join(BASE_DIR, "ARQUIVOS_BASE")
PASTA_SAIDA = os.path.join(BASE_DIR, "ARQUIVOS_ETIQUETA")

# ==========================================================
# FUNÇÕES AUXILIARES
# ==========================================================

def log(msg):
    print(f"[LOG] {msg}")

def limpar_referencia(ref):
    if pd.isna(ref):
        return None
    return re.sub(r"\D", "", str(ref))

def extrair_loja(valor):
    if pd.isna(valor):
        return None
    match = re.search(r"\d+", str(valor))
    return int(match.group()) if match else None

def norm(x):
    return str(x).strip().upper()

# ==========================================================
# LOCALIZAR ARQUIVOS
# ==========================================================

def localizar_arquivos():

    arquivos = os.listdir(PASTA_ENTRADA)

    arquivo_csv = None
    arquivo_xlsx = None

    for arq in arquivos:

        if arq.lower().endswith(".csv"):
            arquivo_csv = arq

        elif arq.lower().endswith(".xlsx"):
            arquivo_xlsx = arq

    if not arquivo_csv:
        raise Exception("Arquivo .csv (modelo OC) nÃ£o encontrado")

    if not arquivo_xlsx:
        raise Exception("Arquivo .xlsx (base de pedidos) nÃ£o encontrado")

    print("\nArquivos encontrados:\n")
    print("Modelo OC (.csv):", arquivo_csv)
    print("Base pedidos (.xlsx):", arquivo_xlsx)

    confirmar = input("\nConfirmar? (y/n): ")

    if confirmar.lower() != "y":
        raise Exception("ConfirmaÃ§Ã£o cancelada")

    return arquivo_csv, arquivo_xlsx


# ==========================================================
# EXECUÇÃO
# ==========================================================

arquivo_oc, arquivo_grade = localizar_arquivos()

log("Arquivos confirmados")

caminho_oc = os.path.join(PASTA_ENTRADA, arquivo_oc)

log("Lendo arquivo OC modelo")

df_oc = pd.read_csv(caminho_oc, sep=";", dtype=str, header=None)

df_oc["REF"] = df_oc.iloc[:,2].str.strip().str.upper()
df_oc["COR"] = df_oc.iloc[:,5].str.strip().str.upper()
df_oc["TAM"] = df_oc.iloc[:,1].astype(str).str.strip()

caminho_grade = os.path.join(PASTA_ENTRADA, arquivo_grade)

hoje = datetime.today().strftime("%d-%m-%Y")

log("Abrindo planilha de pedidos")

wb = load_workbook(caminho_grade, data_only=True)

# SOMENTE ABAS VISÃVEIS
abas_pedidos = [
    s.title for s in wb.worksheets
    if s.title.startswith("PEDIDO_") and s.sheet_state == "visible"
]

log(f"{len(abas_pedidos)} pedidos encontrados")

dados_lojas = {}
total_geral = []
total_pares_geral = 0
total_por_pedido = {}



# ==========================================================
# PROCESSAR PEDIDOS
# ==========================================================

for aba_nome in abas_pedidos:

    log(f"Lendo {aba_nome}")

    ws = wb[aba_nome]

    fornecedor = ws["N2"].value
    marca = ws["N3"].value

    if not fornecedor:
        raise Exception(f"Fornecedor nÃ£o encontrado na aba {aba_nome}")

    fornecedor = str(fornecedor).strip()
    marca = str(marca).strip()

    log(f"Fornecedor: {fornecedor}")
    log(f"Marca: {marca}")

    # ------------------------------------------------------
    # LOJAS
    # ------------------------------------------------------

    lojas = []

    for row in ws.iter_rows(min_row=2, max_row=7, min_col=16, max_col=20):

        for cell in row:

            valor = cell.value

            if valor is None:
                continue

            valor = str(valor).strip()

            match = re.match(r"^\d+", valor)

            if match:
                lojas.append(int(match.group()))

    lojas = sorted(set(lojas))

    log(f"Lojas encontradas: {lojas}")

    # ------------------------------------------------------
    # GRADE A / B / C
    # ------------------------------------------------------

    grade_usada = ws["W9"].value

    linha_grade = {
        "A": 5,
        "B": 6,
        "C": 7
    }[grade_usada]

    tamanhos = []
    colunas_tamanho = []

    for col in range(24, 43): 

        tamanho = ws.cell(row=linha_grade, column=col).value

        if not isinstance(tamanho, (int, float)):
            continue

        # verificar se existe quantidade em algum produto
        tem_quantidade = False

        for linha in range(9, 30):

            qtd = ws.cell(row=linha, column=col).value

            if isinstance(qtd, (int, float)) and qtd > 0:
                tem_quantidade = True
                break

        if tem_quantidade:
            tamanhos.append(int(tamanho))
            colunas_tamanho.append(col)
        
    log(f"Grade usada: {grade_usada}")
    log(f"Tamanhos realmente usados: {tamanhos}")
    log(f"Colunas usadas na grade: {colunas_tamanho}")
    
    # ------------------------------------------------------
    # PRODUTOS
    # ------------------------------------------------------

    for linha in range(9, 30):

        referencia = limpar_referencia(ws.cell(row=linha, column=2).value)
        desc_robo = ws.cell(row=linha, column=3).value
        cor = ws.cell(row=linha, column=4).value
        referencia_final = f"{desc_robo} {referencia}"

        if not referencia:
            continue

        for tamanho, coluna in zip(tamanhos, colunas_tamanho):

            codigo_grade = f"{referencia_final}_{cor}_{tamanho}"

            qtd = ws.cell(row=linha, column=coluna).value

            if qtd is None:
                continue

            try:
                qtd = int(qtd)
            except:
                continue

            if qtd <= 0:
                continue

            match = df_oc[
            (df_oc["REF"] == referencia_final.strip().upper()) &
            (df_oc["COR"] == cor.strip().upper()) &
            (df_oc["TAM"] == str(tamanho))
            ]

            if not match.empty:
                cod_barras = match.iloc[0,0]
                codigo_oc = f"{match.iloc[0]['REF']}_{match.iloc[0]['COR']}_{match.iloc[0]['TAM']}"
            else:
                cod_barras = ""
                codigo_oc = ""
                log(f"CODIGO NÃƒO ENCONTRADO: {referencia_final} | {cor} | {tamanho}")

            log(f"Ref {referencia} | Cor {cor} | Tam {tamanho} | Qtd {qtd}")

            total_pares_geral += qtd * len(lojas)
            total_por_pedido.setdefault(aba_nome, 0)
            total_por_pedido[aba_nome] += qtd * len(lojas)

            match = df_oc[
            (df_oc["REF"] == referencia_final) &
            (df_oc["COR"] == cor) &
            (df_oc["TAM"] == str(tamanho))
            ]

            if not match.empty:
                cod_barras = match.iloc[0,0]
            else:
                cod_barras = ""
                log(f"CODIGO NÃƒO ENCONTRADO: {referencia_final} {cor} {tamanho}")
                    
            
            for loja in lojas:
                for _ in range(qtd):

                    registro = {
                        "COD_BARRAS": cod_barras,
                        "TAMANHO": tamanho,
                        "REFERÃŠNCIA": referencia_final,
                        "MARCA": marca,
                        "COR": cor
                    }

                    total_geral.append(registro)

                    if loja not in dados_lojas:
                        dados_lojas[loja] = []

                    dados_lojas[loja].append(registro)
                    
    log(f"Total de pares lidos em {aba_nome}: {total_por_pedido.get(aba_nome, 0)}")

log("Leitura finalizada")

log(f"TOTAL DE SAPATOS LIDOS: {total_pares_geral}")

# ==========================================================
# CRIAR PASTAS
# ==========================================================

pasta_dia = os.path.join(PASTA_SAIDA, hoje)
pasta_fornecedor = os.path.join(pasta_dia, fornecedor)

os.makedirs(pasta_fornecedor, exist_ok=True)

log("Pastas criadas")

# ==========================================================
# GERAR EXCEL
# ==========================================================

numero_oc = re.search(r"\d+", arquivo_oc).group()

# CONSOLIDADO

df_total = pd.DataFrame(total_geral)

arquivo_total = os.path.join(
    pasta_fornecedor,
    f"OC_{numero_oc}_LJ350.xlsx"
)

df_total.to_excel(arquivo_total, index=False)

log("Arquivo LJ350 criado")

# POR LOJA

for loja, registros in dados_lojas.items():

    df_loja = pd.DataFrame(registros)

    caminho = os.path.join(
        pasta_fornecedor,
        f"OC_{numero_oc}_LJ{loja}.xlsx"
    )

    df_loja.to_excel(caminho, index=False)

    log(f"Arquivo loja {loja} criado")

log("PROCESSO FINALIZADO")
