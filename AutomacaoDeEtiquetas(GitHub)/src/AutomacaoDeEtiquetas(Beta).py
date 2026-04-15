import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import tkinter as tk
from tkinter import scrolledtext
from tkinter import messagebox
import threading

# ==========================================================
# CONFIGURAÃ‡Ã•ES
# ==========================================================

BASE_DIR = os.getenv("AUTOMACAO_ETIQUETAS_BASE_DIR", os.path.dirname(os.path.abspath(__file__)))

PASTA_ENTRADA = os.path.join(BASE_DIR, "ARQUIVOS_BASE")
PASTA_SAIDA = os.path.join(BASE_DIR, "ARQUIVOS_ETIQUETA")

# ==========================================================
# INTERFACE LOG
# ==========================================================

def escrever_log(msg):
    log_area.insert(tk.END, msg + "\n")
    log_area.see(tk.END)
    root.update_idletasks()

# ==========================================================
# FUNÃ‡Ã•ES AUXILIARES
# ==========================================================

def limpar_referencia(ref):
    if pd.isna(ref):
        return None
    return re.sub(r"\D", "", str(ref))

def norm(x):
    return str(x).strip().upper()

# ==========================================================
# LOCALIZAR ARQUIVOS
# ==========================================================

def localizar_arquivos():

    arquivos = os.listdir(PASTA_ENTRADA)

    arquivo_csv = None
    arquivo_xlsx = None

    for arq in arquivos:

        if arq.lower().endswith(".csv"):
            arquivo_csv = arq

        elif arq.lower().endswith(".xlsx"):
            arquivo_xlsx = arq

    if not arquivo_csv:
        raise Exception("Arquivo .csv (modelo OC) nÃ£o encontrado")

    if not arquivo_xlsx:
        raise Exception("Arquivo .xlsx (base pedidos) nÃ£o encontrado")

    escrever_log(f"CSV encontrado: {arquivo_csv}")
    escrever_log(f"XLSX encontrado: {arquivo_xlsx}")

    return arquivo_csv, arquivo_xlsx

def confirmar_arquivos(arquivo_oc, arquivo_grade):

    msg = f"""
Arquivos encontrados:

OC / Modelo:
{arquivo_oc}

Base de pedidos:
{arquivo_grade}

Estes sÃ£o os arquivos corretos?
"""

    resposta = messagebox.askyesno(
        "ConfirmaÃ§Ã£o de arquivos",
        msg
    )

    return resposta
# ==========================================================
# PROCESSO PRINCIPAL
# ==========================================================

def executar():

    try:

        escrever_log("Localizando arquivos...")

        arquivo_oc, arquivo_grade = localizar_arquivos()
        
        if not confirmar_arquivos(arquivo_oc, arquivo_grade):
            escrever_log("Processamento cancelado pelo usuÃ¡rio.")
            return

        caminho_oc = os.path.join(PASTA_ENTRADA, arquivo_oc)

        escrever_log("Lendo OC modelo...")

        df_oc = pd.read_csv(caminho_oc, sep=";", dtype=str, header=None)

        df_oc["REF"] = df_oc.iloc[:,2].str.strip().str.upper()
        df_oc["COR"] = df_oc.iloc[:,5].str.strip().str.upper()
        df_oc["TAM"] = df_oc.iloc[:,1].astype(str).str.strip()

        caminho_grade = os.path.join(PASTA_ENTRADA, arquivo_grade)

        hoje = datetime.today().strftime("%d-%m-%Y")

        escrever_log("Abrindo planilha de pedidos")

        wb = load_workbook(caminho_grade, data_only=True)

        abas_pedidos = [
            s.title for s in wb.worksheets
            if s.title.startswith("PEDIDO_") and s.sheet_state == "visible"
        ]

        escrever_log(f"{len(abas_pedidos)} pedidos encontrados")
        escrever_log(f"{len(abas_pedidos)} pedidos encontrados")

        info_pedidos.set(
        "PEDIDOS: " + ", ".join(abas_pedidos)
            )
        
        dados_fornecedor = {}
        total_geral = []
        total_pares_geral = 0
        total_por_pedido = {}
        lojas_por_pedido = {}
        
        for aba_nome in abas_pedidos:

            escrever_log(f"Lendo {aba_nome}")

            ws = wb[aba_nome]

            fornecedor = str(ws["N2"].value).strip()
            marca = str(ws["N3"].value).strip()

            escrever_log(f"Fornecedor: {fornecedor}")
            escrever_log(f"Marca: {marca}")

            lojas = []

            for row in ws.iter_rows(min_row=2, max_row=7, min_col=16, max_col=20):

                for cell in row:

                    valor = cell.value

                    if valor is None:
                        continue

                    match = re.match(r"^\d+", str(valor))

                    if match:
                        lojas.append(int(match.group()))

            lojas = sorted(set(lojas))

            lojas_por_pedido[aba_nome] = lojas
            total_por_pedido[aba_nome] = 0

            escrever_log(f"Lojas encontradas: {lojas}")

            grade_usada = str(ws["W9"].value).strip().upper()

            linha_grade = {
                "A":5,
                "B":6,
                "C":7
            }[grade_usada]

            tamanhos = []
            colunas_tamanho = []

            for col in range(24,43):

                tamanho = ws.cell(row=linha_grade,column=col).value

                if not isinstance(tamanho,(int,float)):
                    continue

                tem_qtd = False

                for linha in range(9,30):

                    qtd = ws.cell(row=linha,column=col).value

                    if isinstance(qtd,(int,float)) and qtd>0:
                        tem_qtd=True
                        break

                if tem_qtd:
                    tamanhos.append(int(tamanho))
                    colunas_tamanho.append(col)

            escrever_log(f"Grade {grade_usada}")
            escrever_log(f"Tamanhos usados {tamanhos}")

            for linha in range(9,30):

                referencia = limpar_referencia(ws.cell(row=linha,column=2).value)
                desc_robo = ws.cell(row=linha,column=3).value
                cor = str(ws.cell(row=linha,column=4).value).strip().upper()

                referencia_final = f"{desc_robo} {referencia}"

                if not referencia:
                    continue

                for tamanho,coluna in zip(tamanhos,colunas_tamanho):

                    qtd = ws.cell(row=linha,column=coluna).value

                    if qtd is None:
                        continue

                    try:
                        qtd = int(qtd)
                    except:
                        continue

                    if qtd<=0:
                        continue

                    match = df_oc[
                        (df_oc["REF"]==referencia_final.strip().upper()) &
                        (df_oc["COR"]==cor.strip().upper()) &
                        (df_oc["TAM"]==str(tamanho))
                    ]

                    if not match.empty:
                        cod_barras = match.iloc[0,0]
                    else:
                        cod_barras = ""
                        escrever_log(f"CODIGO NÃƒO ENCONTRADO {referencia_final} {cor} {tamanho}")

                    escrever_log(f"{referencia_final} | {cor} | {tamanho} | Qtd {qtd}")

                    total_pares_geral += qtd*len(lojas)

                    total_por_pedido[aba_nome] += qtd * len(lojas)

                    for loja in lojas:
                        for _ in range(qtd):

                            registro={
                                "COD_BARRAS":cod_barras,
                                "TAMANHO":tamanho,
                                "REFERÃŠNCIA":referencia_final,
                                "MARCA":marca,
                                "COR":cor
                            }

                            total_geral.append(registro)

                            if fornecedor not in dados_fornecedor:
                                dados_fornecedor[fornecedor] = {}

                            if loja not in dados_fornecedor[fornecedor]:
                                dados_fornecedor[fornecedor][loja] = []

                            dados_fornecedor[fornecedor][loja].append(registro)

        linhas_resumo = []

        for pedido in abas_pedidos:

            lojas = lojas_por_pedido.get(pedido, [])
            pares = total_por_pedido.get(pedido, 0)

            linhas_resumo.append(f"{pedido} â†’ Lojas {lojas} â†’ {pares} pares")

        texto_resumo = "\n".join(linhas_resumo)

        info_pedidos.set(texto_resumo)
        info_total.set(f"TOTAL DE SAPATOS: {total_pares_geral}")

        escrever_log(f"TOTAL DE SAPATOS {total_pares_geral}")

        pasta_dia = os.path.join(PASTA_SAIDA, hoje)

        os.makedirs(pasta_dia, exist_ok=True)

        numero_oc = re.search(r"\d+", arquivo_oc).group()

        for fornecedor, lojas in dados_fornecedor.items():

            pasta_fornecedor = os.path.join(pasta_dia, fornecedor)

            contador = 1
            pasta_base = pasta_fornecedor

            while os.path.exists(pasta_fornecedor):
                pasta_fornecedor = f"{pasta_base}_{contador}"
                contador += 1

            os.makedirs(pasta_fornecedor)

            total_fornecedor = []

            for loja, registros in lojas.items():

                total_fornecedor.extend(registros)

                df_loja = pd.DataFrame(registros)

                caminho = os.path.join(
                    pasta_fornecedor,
                    f"OC_{numero_oc}_LJ{loja}.xlsx"
                )

                df_loja.to_excel(caminho, index=False)

                escrever_log(f"{fornecedor} - Arquivo Loja {loja} Criado")

            df_total = pd.DataFrame(total_fornecedor)

            caminho_total = os.path.join(
                pasta_fornecedor,
                f"OC_{numero_oc}_LJ350.xlsx"
            )

            df_total.to_excel(caminho_total, index=False)

            escrever_log(f"{fornecedor} - Arquivo Consolidado Criado")

        escrever_log("PROCESSO FINALIZADO")

    except Exception as e:
        escrever_log(f"ERRO: {e}")


# ==========================================================
# THREAD
# ==========================================================

def verificar_thread(thread):

    if thread.is_alive():
        root.after(100, verificar_thread, thread)
    else:
        mostrar_botoes_finais()

def iniciar():

    btn.config(text="PROCESSANDO...", state="disabled")

    t = threading.Thread(target=executar)
    t.start()

    root.after(100, verificar_thread, t)

def reiniciar():

    log_area.delete("1.0", tk.END)

    info_pedidos.set("Pedidos: -")
    info_total.set("Total de sapatos: -")

    btn_reiniciar.pack_forget()
    btn_encerrar.pack_forget()

    btn.config(text="INICIAR PROCESSAMENTO", state="normal")
    btn.pack(side="left", padx=5)

def encerrar():
    root.destroy()

def mostrar_botoes_finais():

    btn.pack_forget()

    btn_reiniciar.pack(side="left", padx=5)
    btn_encerrar.pack(side="left", padx=5)

    root.update_idletasks()

# ==========================================================
# INTERFACE
# ==========================================================

root=tk.Tk()
root.title("AutomaÃ§Ã£o de Etiquetas")
root.geometry("900x600")

info_pedidos = tk.StringVar(root)
info_total = tk.StringVar(root)

info_pedidos.set("Pedidos: -")
info_total.set("Total de sapatos: -")

titulo=tk.Label(root,text="AUTOMAÃ‡ÃƒO DE ETIQUETAS",font=("Arial",16,"bold"))
titulo.pack(pady=10)

frame_botoes = tk.Frame(root)
frame_botoes.pack(pady=10)

btn = tk.Button(
    frame_botoes,
    text="INICIAR PROCESSAMENTO",
    font=("Arial",12),
    command=iniciar
)
btn.pack()

btn_reiniciar = tk.Button(
    frame_botoes,
    text="REINICIAR PROCESSAMENTO",
    font=("Arial",12),
    command=reiniciar
)

btn_encerrar = tk.Button(
    frame_botoes,
    text="ENCERRAR APLICAÃ‡ÃƒO",
    font=("Arial",12),
    command=encerrar
)

frame_info = tk.Frame(root)
frame_info.pack(pady=5)

label_pedidos = tk.Label(
    frame_info,
    textvariable=info_pedidos,
    font=("Arial",11,"bold")
)
label_pedidos.pack()

label_total = tk.Label(
    frame_info,
    textvariable=info_total,
    font=("Arial",11,"bold")
)
label_total.pack()

log_area=scrolledtext.ScrolledText(root,width=110,height=30)
log_area.pack(padx=10,pady=10)

root.mainloop()
